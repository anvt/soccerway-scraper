from openpyxl import load_workbook


def spread(match_dict, spreadsheet_path):
    """
    Write match data to spreadsheet.
    :param match_dict: Dictionary containing match data
    :param spreadsheet_path: Path to spreadsheet data is to be written to.
    :return:
    """
    wb = load_workbook(spreadsheet_path)
    ws = wb.worksheets[0]
    row = ws.max_row + 1

    # integers in main_cols and time_cols correspond to spreadsheet column numbers.
    main_cols = {
        'week': 1,
        'date': 2,
        'home_team_name': 3,
        'away_team_name': 4,
        'home_goal_total': 5,
        'away_goal_total': 6,
        'referee': 47,
        'home_corners': 48,
        'away_corners': 49,
        'home_shots_on': 54,
        'away_shots_on': 56,
        'home_shots_wide': 55,
        'away_shots_wide': 57,
        'home_fouls': 52,
        'away_fouls': 53,
        'home_offsides': 50,
        'away_offsides': 51,
    }

    time_cols = {
        'home_goal_times': 7,
        'away_goal_times': 15,
        'home_yellow_times': 23,
        'away_yellow_times': 32,
        'home_red_times': 41,
        'away_red_times': 44
    }

    # Write main_cols markets to sheet.
    for stat, col in main_cols.items():
        ws.cell(row=row, column=col).value = match_dict[stat]

    # Write time_cols markets to sheet.
    for stat, start_col in time_cols.items():
        for match_time in match_dict[stat]:
            ws.cell(row=row, column=start_col).value = match_time
            start_col += 1

    wb.save(spreadsheet_path)
    print(f"{match_dict['date']} - {match_dict['home_team_name']} vs {match_dict['away_team_name']} added.")
